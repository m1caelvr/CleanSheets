import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from src.sheet_selection import select_sheet
from src.column_selection import select_columns
from src.file_operations import remove_columns

def show_excluded_columns(columns_to_remove):
    if columns_to_remove:
        print(f"Colunas selecionadas para exclusão: {columns_to_remove}")
    else:
        print("Nenhuma coluna foi selecionada para remoção.")

def select_file_and_remove_columns():
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if file_path:
        workbook = load_workbook(filename=file_path)
        sheet_name = select_sheet(workbook)

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            columns_to_remove = select_columns(sheet, [])

            show_excluded_columns(columns_to_remove)

            if columns_to_remove:
                remove_columns(file_path, sheet_name, columns_to_remove)
                print(f"As colunas {columns_to_remove} foram removidas da planilha {sheet_name} no arquivo {file_path}")
            else:
                print("Nenhuma coluna foi selecionada para remoção.")
        else:
            print(f"A planilha '{sheet_name}' não existe no arquivo.")
