import pandas as pd
from openpyxl import load_workbook

def file_treatment(file):
    try:
        xls = pd.ExcelFile(file.path)
        sheetnames = xls.sheet_names

        wb = load_workbook(file.path, read_only=True)
        column_counts = {sheet: wb[sheet].max_column for sheet in sheetnames}
        row_counts = {sheet: wb[sheet].max_row for sheet in sheetnames}
        wb.close()

        # print(f'rows: {row_counts} / {column_counts}')
        
        return sheetnames, column_counts
    except Exception as e:
        print(f"Erro ao processar arquivo: {e}")
        return None, None

def get_columns_from_sheet(file, sheet_name, treatment, specified):
    try:
        if treatment == 'automatic':
            xls = pd.ExcelFile(file.path)
            df = xls.parse(sheet_name, engine='openpyxl', header=None)

            column_names = df.head(10).apply(lambda col: col.dropna().iloc[0], axis=0).tolist()
            
        else:
            specified = int(specified)
            df = pd.read_excel(file.path, sheet_name=sheet_name, header=None, nrows=specified + 10)

            column_names = df.iloc[specified - 1].dropna().tolist()

        return column_names if column_names else []

    except Exception as e:
        print(f"Erro ao processar arquivo: {e}")
        return []
