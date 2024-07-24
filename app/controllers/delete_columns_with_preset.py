import pandas as pd
from openpyxl import load_workbook
from app.utils.get_data_json import load_json_data

def remove_hidden_rows_and_columns(df, ws):
    hidden_cols = set()

    rows_quantity = len(ws['A'])
    if rows_quantity > 1510:
        rows_quantity = 1510
    
    column_quantity = 0
    for col in ws.column_dimensions:
        column_quantity += 1
        if ws.column_dimensions[col].hidden:
            hidden_cols.add(ws[col + '1'].column - 1)

    visible_cols = df.drop(df.columns[list(hidden_cols)], axis=1, errors='ignore')

    return visible_cols, rows_quantity, column_quantity

def delete_columns_with_preset(path, sheet, preset_name, keeps):
    try:
        presets = load_json_data()
                
        if preset_name not in presets:
            print(f"Preset '{preset_name}' não encontrado no JSON.")
            return f"Preset '{preset_name}' não encontrado no JSON.", [], ""
        
        preset_data = presets[preset_name]
        preset_columns = set(preset_data.get('Columns', []))
        
        print(f'preset_columns: {len(preset_columns)}')
        
        if not preset_columns:
            print(f"Nenhuma coluna especificada para o preset '{preset_name}'.")
            return f"Nenhuma coluna especificada para o preset '{preset_name}'.", [], ""

        wb = load_workbook(filename=path)
        ws = wb[sheet]
        
        sample_df = pd.read_excel(path, sheet_name=sheet, nrows=100)

        total_cols = sample_df.shape[1]

        df, rows_quantity, column_quantity = remove_hidden_rows_and_columns(sample_df, ws)
                
        print(f"O arquivo tem mais de {rows_quantity} linhas e mais de {column_quantity} / {total_cols} colunas na amostra.")

        if rows_quantity > 1500 or total_cols > 50:
            print("Usando pandas para manipulação de dados.")
            return process_with_pandas(path, sheet, preset_columns, keeps, ws)
        else:
            print("Usando openpyxl para manipulação de dados.")
            return process_with_openpyxl(path, sheet, preset_columns, keeps)
        
    except Exception as e:
        print(f"Erro ao processar: {e}")
        return f"Erro ao processar: {e}", [], ""

def process_with_pandas(path, sheet, preset_columns, keeps, ws):
    try:
        df = pd.read_excel(path, sheet_name=sheet)

        df, rows_quantity, column_quantity = remove_hidden_rows_and_columns(df, ws)

        cols_to_remove = []
        cols_not_found = [] 
        for col in df.columns:
            found_in_preset = False
            not_found_values = []
            
            for value in df[col].head(10).unique():
                if value in preset_columns:
                    found_in_preset = True
                    break
                else:
                    not_found_values.append(value)
            
            if keeps == 'true':
                if not found_in_preset:
                    cols_to_remove.append(col)
                else:
                    cols_not_found.append((col, not_found_values))
            else:
                if found_in_preset:
                    cols_to_remove.append(col)
                else:
                    cols_not_found.append((col, not_found_values))
                
        if not cols_to_remove:
            action = 'manter' if keeps == 'true' else 'deletar'
            print(f"Nenhuma coluna encontrada para {action} no preset.")
            return f"Nenhuma coluna encontrada para {action} no preset.", [], ""

        remaining_columns = [col for col in df.columns if col not in cols_to_remove]

        df = df[remaining_columns]

        with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
        
        action = 'mantidas' if keeps == 'true' else 'deletadas'
        print(f"{len(cols_to_remove)} Colunas {action} com sucesso do arquivo {path} na planilha {sheet}.")
        status_msg = f"Colunas {action} com sucesso do arquivo {path} na planilha {sheet}."
        found_msg = f"Foram {action} {len(preset_columns)} colunas do preset na planilha."
        return status_msg, list(cols_to_remove), found_msg

    except Exception as e:
        print(f"Erro ao processar com pandas: {e}")
        return f"Erro ao processar com pandas: {e}", [], ""

def process_with_openpyxl(path, sheet, preset_columns, keeps):
    try:
        wb = load_workbook(filename=path)
        ws = wb[sheet]

        cols_to_remove = []
        cols_not_found = []

        for col in ws.iter_cols(min_row=2, max_row=11):
            col_letter = col[0].column_letter
            found_in_preset = False
            not_found_values = []

            for cell in col:
                if cell.value in preset_columns:
                    found_in_preset = True
                    break
                else:
                    not_found_values.append(cell.value)
            
            if keeps == 'true':
                if not found_in_preset:
                    cols_to_remove.append(col[0].column)
                else:
                    cols_not_found.append((col_letter, not_found_values))
            else:
                if found_in_preset:
                    cols_to_remove.append(col[0].column)
                else:
                    cols_not_found.append((col_letter, not_found_values))

        if not cols_to_remove:
            action = 'manter' if keeps == 'true' else 'deletar'
            print(f"Nenhuma coluna encontrada para {action} no preset.")
            return f"Nenhuma coluna encontrada para {action} no preset.", [], ""

        for col_index in reversed(sorted(cols_to_remove)):
            ws.delete_cols(col_index)
        
        wb.save(path)
        
        action = 'mantidas' if keeps == 'true' else 'deletadas'
        print(f"{len(cols_to_remove)} Colunas {action} com sucesso do arquivo {path} na planilha {sheet}.")
        status_msg = f"Colunas {action} com sucesso do arquivo {path} na planilha {sheet}."
        found_msg = f"Foram {action} {len(preset_columns)} colunas do preset na planilha."
        return status_msg, list(cols_to_remove), found_msg

    except Exception as e:
        print(f"Erro ao processar com openpyxl: {e}")
        return f"Erro ao processar com openpyxl: {e}", [], ""
