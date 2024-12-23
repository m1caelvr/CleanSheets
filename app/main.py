import flet as ft
import asyncio
import openpyxl as op
import updates.check_for_update as update
import app.controllers.file_treatment as ftm
from functools import partial
from app.utils.get_data_json import load_json_data
from app.utils.add_new_preset import add_new_preset
from app.controllers.delete_columns import delete_columns
from app.utils.get_presets_names import get_presets_names
from app.utils.add_column_to_json import add_column_to_json
from app.data.json_handler import ensure_documents_json_file
from app.utils.remove_preset_from_json import remove_preset_from_json
from app.utils.remove_column_from_json import remove_column_from_json
from app.share_presets.import_preset_exported import import_preset_exported
from app.share_presets.create_file_to_export import create_file_to_export
from app.controllers.delete_columns_with_preset import delete_columns_with_preset
from app.utils.update_preset_in_json import update_preset_exist_in_json, update_preset_in_json 

def main(page: ft.Page):
    initial_width = 550
    initial_height = 600
    # page.window.always_on_top = True
    page.window.width = initial_width
    page.window.height = initial_height
    page.title = 'CleanSheets'
    page.scroll = 'always'
    page.horizontal_alignment = 'center'

    new_version, local_version, remote_version = update.check_for_update()
    # print(f'new version: {'yes' if new_version == True else 'no'} / {remote_version}')

    def download_new_version(e):
        update.download_and_apply_update()

    action_button_style = ft.ButtonStyle(color=ft.colors.BLUE)
    banner_new_vesion = ft.Banner(
        bgcolor=ft.colors.PRIMARY_CONTAINER,
        leading=ft.Icon(ft.icons.WARNING_AMBER_ROUNDED, size=40),
        content=ft.Text(
            value=f"Uma nova versão do app está disponível para baixar.\n{local_version} -> {remote_version}",
            weight="bold",
            color=ft.colors.ON_PRIMARY_CONTAINER,
        ),
        actions=[
            ft.TextButton(text="Lembre mais tarde", on_click=lambda _: page.close(banner_new_vesion)),
            ft.TextButton(text="Baixar", on_click=download_new_version)
        ],
    )

    
    title_value = ft.Text(
        size=28,
        weight="bold",
        value='CleanSheets',
    )
    title = ft.Row(
        controls=[title_value],
        alignment='center',
    )

    wrapper_title_value = ft.Text(
        value='Selecione o arquivo',
        size=20,
    )
    
    checkboxes_row = ft.Row(
        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
        width=initial_width,
        spacing=10,
        wrap=True,
    )
    
    checkboxes = []
    selected_sheet = None
    selected_file = None
    selected_treatment = "automatic"
    line_specified = 1
    value_line = 0

    def update_selected_columns():
        selected_columns = [cb.content.label for cb in checkboxes if cb.content.value]
        if selected_columns:
            selected_columns_text.value = f'{len(selected_columns)} colunas selecionadas.'
            selected_columns_container.visible = True
        else:
            selected_columns_container.visible = False

        page.update()

    def inputs_create(column_names):
        checkboxes.clear()
        for i, column_name in enumerate(column_names):
            checkbox = ft.Checkbox(
                label=f'{i+1} - {column_name}',
                value=False,
                width=210,
                height=30,
                on_change=lambda _: update_selected_columns()
            )
            checkbox_container = ft.Container(
                content=checkbox,
                padding=ft.padding.all(5),
                border=ft.border.all(1, ft.colors.OUTLINE_VARIANT),
                border_radius=ft.border_radius.all(7),
            )
            checkboxes.append(checkbox_container)
        checkboxes_row.controls = checkboxes
        page.update()
        print('inputs created')

    def handle_close(e):
        nonlocal selected_sheet, selected_file, selected_treatment, line_specified

        button_clicked = e.control.text

        if button_clicked == 'Cancelar':
            page.close(dlg_modal)

            return
        
        if not number_input.value and selected_treatment == "manual":
            number_input.error_text = '*'
            page.update()
                
            return

        if selected_sheet and selected_file:
            _, column_counts = ftm.file_treatment(selected_file)
            
            if selected_sheet in column_counts:
                line_specified = number_input.value

                selected_columns = ftm.get_columns_from_sheet(selected_file, selected_sheet, selected_treatment, line_specified)
                
                inputs_create(selected_columns)
                
                result_columns.value = f'Número de colunas: {len(selected_columns)}'
                
                page.close(dlg_modal)
                page.update()
            else:
                print(f"Planilha '{selected_sheet}' não encontrada nos dados de contagem de colunas.")
        else:
            print("Selecione um arquivo e uma planilha antes de continuar.")

    width_input = 190

    number_input = ft.TextField(
        keyboard_type=ft.KeyboardType.NUMBER,
        label="Exemplo: '1' ",
        width=width_input,
        border_color=ft.colors.OUTLINE_VARIANT,
        visible=False,
    )

    dlg_modal = ft.AlertDialog(
        modal=True,
        title=ft.Row(
            controls=[ft.Text("Planilhas Disponíveis", size=20)],
            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            wrap=True,
        ),
        actions=[
            ft.Row(
                controls=[
                    ft.TextButton("Cancelar", on_click=handle_close),
                    ft.TextButton("Seguir", on_click=handle_close),
                ],
                width=initial_width,
                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            ),
            
        ],
    )
            
    def toggle_number_input(i):
        nonlocal number_input
        if i == 1:
            number_input.visible = True
            page.update()
        else:
            number_input.visible = False
            page.update()
        page.update()

    def on_dropdown_change(e):
        nonlocal selected_treatment, value_line, number_input
        selected_treatment = e.control.value
        toggle_number_input(1 if selected_treatment == "manual" else 0)

        value_line = number_input.value if selected_treatment == "manual" else 1

        page.update()

    dropdown_and_number_input_row = ft.Dropdown(
        label="Selecionar primeira linha",
        options=[
            ft.dropdown.Option(
                text="Automático",
                key="automatic",
            ),
            ft.dropdown.Option(
                text="Manual",
                key="manual",
            ),
        ],
        value="automatic",
        width=width_input,
        border_color=ft.colors.OUTLINE_VARIANT,
        text_size=14,
        on_change=on_dropdown_change,
    )
    
    def on_result(e: ft.FilePickerResultEvent):
        selected_columns_container.visible = False
        result_action.visible = False

        SUPPORTED_EXTENSIONS = ['xlsx', 'xls']
        nonlocal selected_file

        if e.files:
            selected_file = e.files[0]
            file_name_value = selected_file.name
            file_path = selected_file.path
            ext = file_name_value.split('.')[-1].lower()

            if ext in SUPPORTED_EXTENSIONS:
                sheetnames, columns_sheet = ftm.file_treatment(selected_file)
                
                def on_container_click(index):
                    page.update()
                    
                    nonlocal selected_sheet
                    selected_sheet = sheetnames[index]

                    for i, container in enumerate(radio_containers):
                        if i == index:
                            container.bgcolor = ft.colors.OUTLINE_VARIANT
                        else:
                            container.bgcolor = ft.colors.TRANSPARENT
                        
                    page.update()

                radio_containers = [
                    ft.Container(
                        content=ft.Row(
                            controls=[
                                ft.Text(
                                    value=f"Planilha: {sheet}",
                                    size=14,
                                ),
                                ft.Text(
                                    value=f"Colunas: {columns_sheet.get(sheet, 0)}",
                                    size=11,
                                )
                            ],
                            spacing=5,
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        ),
                        border=ft.border.all(1, ft.colors.OUTLINE_VARIANT),
                        border_radius=ft.border_radius.all(7),
                        padding=ft.padding.all(10),
                        width=initial_width,
                        on_click=lambda _, index=i: on_container_click(index),
                    )
                    for i, sheet in enumerate(sheetnames)
                ]

                radio_group = ft.Row(
                    controls=[
                        ft.Row(
                            controls=[
                                dropdown_and_number_input_row,
                                number_input,
                            ],
                            wrap=True,
                            width=initial_width,
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            vertical_alignment=ft.CrossAxisAlignment.START,
                        ),
                        *radio_containers,
                    ],
                    spacing=10,
                    width=initial_width,
                    wrap=True,
                )

                nonlocal dlg_modal, checkboxes

                dlg_modal.content = radio_group
                page.open(dlg_modal)

                result_container.visible = True
                checkboxes.clear()
                
                file_path_text.value = f"Caminho: {file_path}"
                file_name.value = f"Arquivo: {file_name_value}"

                if sheetnames is None:
                    result_title.value = 'Erro:'
                    file_path_text.value = 'Erro ao processar arquivo.'
            else:
                result_title.value = 'Erro:'
                file_path_text.value = 'Apenas arquivos de planilha (Excel) são permitidos.'

            page.update()
    
    file_name = ft.Text()
    result_columns = ft.Text()
    
    result_title = ft.Text(size=16)
    file_path_text = ft.Text()

    file_picker = ft.FilePicker(on_result=on_result)
    page.overlay.append(file_picker)

    pick_file_button = ft.ElevatedButton(
        on_click=lambda _: file_picker.pick_files(),
        width=initial_width,
        text="Upload",
    )
    
    returned_mensagem = ft.Text()
    
    banner_returned_import = ft.Banner(
        bgcolor=ft.colors.OUTLINE_VARIANT,
        content=returned_mensagem,
        actions=[
            ft.TextButton(text="Fechar", on_click=lambda _: page.close(banner_returned_import)),
        ],
    )

    def import_preset(e: ft.FilePickerResultEvent):
        try:
            file_with_preset = e.files[0].path
            if file_with_preset:
                feedback_mensage = import_preset_exported(file_with_preset)
        except Exception:
            feedback_mensage = 'Nenhum arquivo selecionado.'

        returned_mensagem.value = feedback_mensage
        page.open(banner_returned_import)
        
        create_presets_names()
        page.update()

    file_imported = ft.FilePicker(on_result=import_preset)
    page.overlay.append(file_imported)

    def export_preset(e):
        presets_to_export = presets_to_export_group
        create_file_to_export(presets_to_export)
        page.close(export_preset_modal)

    presets_to_export_area = ft.Row(wrap=True, visible=False)

    def presets_to_export():
        global presets_to_export_group

        presets_to_export_group = []

        def add_preset_to_export_group(e):
            preset_selected = e.control.label

            if e.control.value:
                if preset_selected not in presets_to_export_group:
                    presets_to_export_group.append(preset_selected)
            else:
                if preset_selected in presets_to_export_group:
                    presets_to_export_group.remove(preset_selected)

            export_preset_button.disabled = False if presets_to_export_group else True

            print(f'presets to export: {presets_to_export_group}')
            page.update()

        presets_to_export_elements = []
        presets_value = get_presets_names()
        
        for i, preset in enumerate(presets_value):
            checkbox = ft.Checkbox(
                label=f'{preset}',
                value=False,
                width=195,
                height=30,
                on_change=add_preset_to_export_group,
            )
            preset_container = ft.Container(
                content=checkbox,
                padding=ft.padding.all(5),
                border=ft.border.all(1, ft.colors.OUTLINE_VARIANT),
                border_radius=ft.border_radius.all(7),
            )
            presets_to_export_elements.append(preset_container)
        presets_to_export_area.controls = presets_to_export_elements
        presets_to_export_area.visible = True
        page.update()
        
    export_preset_button = ft.TextButton("Exportar", on_click=export_preset, disabled=True)

    def def_all_presets_to_export():
        global presets_to_export_group

        presets_to_export_group = []
        presets_group = get_presets_names()

        for i, preset in enumerate(presets_group):
            presets_to_export_group.append(preset)

    def def_preset_to_export(e):
        options_changed_value = e.control.value

        if options_changed_value == 'all':
            def_all_presets_to_export()
            export_preset_button.disabled = False
            presets_to_export_area.visible = False
        else:
            presets_to_export()
            export_preset_button.disabled = True

        page.update()

    export_preset_modal = ft.AlertDialog(
        modal=True,
        title=ft.Row(
            controls=[
                ft.Text("Exportar presets", size=20),
            ],
            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            wrap=True,
        ),
        content=ft.Row(
            controls=[
                ft.RadioGroup(
                    content=ft.Row(
                        controls=[
                            ft.Radio(value="all", label="Exportar todos", key='all'),
                            ft.Radio(value="especifics", label="Personalizar"),
                        ]
                    ),
                    value='all',
                    on_change=def_preset_to_export,
                ),
                presets_to_export_area
            ],
            wrap=True,
        ),
        actions=[
            ft.Row(
                controls=[
                    ft.TextButton("Cancelar", on_click=lambda _: page.close(export_preset_modal)),
                    export_preset_button,
                ],
                width=initial_width,
                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            ),
        ],
    )

    def open_export_preset(e):
        def_all_presets_to_export()
        
        export_preset_button.disabled = False

        page.open(export_preset_modal)
        page.update()

    import_and_export = ft.Container(
        content=ft.Row(
            controls=[
                ft.TextButton(
                    icon=ft.icons.DRIVE_FOLDER_UPLOAD_OUTLINED,
                    text="Importar preset",
                    on_click=lambda _: file_imported.pick_files(),
                ),
                ft.TextButton(
                    icon=ft.icons.SEND_TIME_EXTENSION_OUTLINED,
                    text="Export preset",
                    on_click=open_export_preset,
                ),
            ],
        ),
        width=initial_width,
    )

    def on_radio_change(e):
        delete_button.disabled = False
        page.update()

    radio_group_columns = ft.RadioGroup(
        content=ft.Row(
            controls=[
                ft.Radio(value="delete", label="Excluir colunas"),
                ft.Radio(value="keep", label="Manter colunas"),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
        ),
        on_change=on_radio_change,
    )

    result_action_text = ft.Row(
        alignment=ft.MainAxisAlignment.CENTER,
        expand=True,
        wrap=True,
    )
    result_action = ft.Container(
        alignment=ft.alignment.center,
        content=result_action_text,
        visible=False,
    )
    
    def get_action_message(action, column_count):
        if column_count > 1:
            return "mantidas" if action == "keep" else "deletadas"
        else:
            return "mantida" if action == "keep" else "deletada"
    
    def clear_page(action, selected_columns):
        action_trated = get_action_message(action, len(selected_columns))
        columns_text = 'Colunas' if len(selected_columns) > 1 else 'Coluna'
                
        result_action_text.controls = [
            ft.Text(f"{columns_text}: {selected_columns} {action_trated} com sucesso!", size=16, weight="bold"),
        ]
        result_action.visible = True
        result_container.visible = False
        progress_deleting.visible = False
        
        page.update()

    def columns_to_delete():
        delete_button.disabled = True
        progress_deleting.visible = True
        page.update()

        selected_columns = [cb.content.label.split(' - ')[0] for cb in checkboxes if cb.content.value]
        selected_columns = list(map(int, selected_columns))
        if selected_columns:
            action = radio_group_columns.value
            file_path = selected_file.path
            sheet_name = selected_sheet
            
            if action == "delete":
                delete_columns(file_path, sheet_name, selected_columns)
            else:
                workbook = op.load_workbook(filename=file_path)
                sheet = workbook[sheet_name]
                all_columns = list(range(1, sheet.max_column + 1))
                columns_to_keep = [col for col in all_columns if col not in selected_columns]
                delete_columns(file_path, sheet_name, columns_to_keep)
                workbook.close()

            clear_page(action, selected_columns)
        else:
            print("Nenhuma coluna selecionada")

    delete_button = ft.ElevatedButton(
        on_click=lambda _: columns_to_delete(),
        text="Deletar colunas",
        disabled=True,
    )

    progress_deleting = ft.Container(
        margin=ft.Margin(top=15, bottom=0, left=0, right=0),
        visible=False,
        content=ft.ProgressBar(
            border_radius=ft.border_radius.all(0),
            bgcolor=ft.colors.TRANSPARENT,
            width=initial_width,
            color=ft.colors.OUTLINE_VARIANT,
        )
    )

    selected_columns_text = ft.Text(text_align=ft.TextAlign.CENTER)
    selected_columns_container = ft.Container(
        content=ft.Column(
            width=initial_width,
            controls=[
                selected_columns_text,
                radio_group_columns,
                delete_button,
                progress_deleting,
            ],
            alignment=ft.MainAxisAlignment.START,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        visible=False,
        width=initial_width,
        padding=ft.padding.all(10),
        border=ft.border.all(1, ft.colors.OUTLINE_VARIANT),
        border_radius=ft.border_radius.all(7),
    )

    result_container = ft.Container(
        width=initial_width,
        content=ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        file_name,
                        result_columns,
                    ],
                    spacing=10,
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    wrap=True,
                    width=initial_width,
                ),
                file_path_text,
                checkboxes_row,
                result_title,
                selected_columns_container,
            ]
        )
    )

    home_area_route = ft.Container(
        visible=True,
        padding=ft.margin.all(20),
        width=initial_width,
        content=ft.Column(
            controls=[
                ft.Row(
                    alignment=ft.MainAxisAlignment.CENTER,
                    controls=[wrapper_title_value]
                ),
                pick_file_button,
                result_container,
                result_action,
            ],
        ),
    )
    
    def column_to_storage(e):
        nonlocal column_to_storage_input

        button_clicked = e.control.text

        if button_clicked == 'Cancelar':
            page.close(add_column_modal)

            return
        else:
            if not column_to_storage_input.value:
                
                column_to_storage_input.error_text = 'Digite o nome da coluna'
                page.update()
                
                return
            else:
                column_to_storage_input.error_text = None
                page.update()

            column_to_storage_value = column_to_storage_input.value

            add_column_to_json(current_preset_column, column_to_storage_value)
            create_json_elements(False, current_preset_column)
            
            page.close(add_column_modal)
            page.update()
    
    column_to_storage_input = ft.TextField(
        hint_text="Digite o nome da coluna",
        border_color=ft.colors.OUTLINE_VARIANT,
    )
    
    add_column_modal = ft.AlertDialog(
        modal=True,
        title=ft.Row(
            controls=[
                ft.Text("Coluna para armazenar", size=20),
                column_to_storage_input,
            ],
            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            wrap=True,
        ),
        actions=[
            ft.Row(
                controls=[
                    ft.TextButton("Cancelar", on_click=column_to_storage),
                    ft.TextButton("Adicionar", on_click=column_to_storage),
                ],
                width=initial_width,
                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            ),
        ],
    )
    
    def open_modal(e):
        page.open(add_column_modal)
        column_to_storage_input.value = ''
        page.update()

    EQS_title_value = ft.Text(
        value='Presets área',
        size=20,
    )

    edit_preset = ft.Container(
        border=ft.border.all(1, ft.colors.OUTLINE_VARIANT),
        border_radius=ft.border_radius.all(10),
        padding=ft.padding.all(12),
        margin=ft.Margin(left=0, top=0, right=0, bottom=10),
        on_click=open_modal,
        content=ft.Row(
            alignment=ft.MainAxisAlignment.CENTER,
            controls=[
                ft.Icon(ft.icons.ADD, size=20),
                ft.Text('Adicionar coluna')
            ]
        )
    )
    
    def disabled_delete_preset(e):
        value_agree_button = agree_box.value
            
        delete_preset_button_agree.disabled = False if value_agree_button == True else True
        delete_preset_button_agree.opacity = 1 if value_agree_button == True else 0.5
        
        page.update()
        
    def delete_preset(e):
        initial_value = current_preset
        
        remove_preset_from_json(initial_value)
        
        page.close(edit_preset_modal)
        create_presets_names()
        page.update()
    
    agree_box = ft.Checkbox(
        label='Li e concordo.',
        value=True,
        on_change=disabled_delete_preset,
    )
    
    delete_preset_button_agree = ft.Container(
        content=ft.Text(
            "Deletar",
            color=ft.colors.RED_400,
            weight=ft.FontWeight.W_500,
        ),
        key='delete',
        expand=1,
        padding=ft.padding.all(10),
        border_radius=ft.border_radius.all(10),
        border=ft.border.all(1, ft.colors.RED_400),
        on_click=delete_preset,
        alignment=ft.alignment.center,
    )
    
    confirm_action = [
        ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        delete_preset_button_agree
                    ],
                    width=initial_width,
                ),
                ft.TextButton("Cancelar", width=initial_width, on_click=lambda _: page.close(edit_preset_modal)),
            ],
            width=initial_width,
        ),
    ]
    
    def confirm_action_to_del_preset():
        initial_value = current_preset
        print(initial_value)
        
        edit_preset_modal.content = ft.Row(
            controls=[
                ft.Text(
                    f'Ciente que ao clicar em deletar, estarei removendo o preset "{initial_value}" e suas colunas existentes. essa ação será irrevessível.',
                    text_align=ft.TextAlign.CENTER,
                    size=12,
                ),
                ft.Row(
                    controls=[agree_box],
                    alignment=ft.MainAxisAlignment.CENTER,
                ),
            ],
            width=initial_width,
            wrap=True
        )
        
        edit_preset_modal.actions = confirm_action
        disabled_delete_preset(None)
        page.update()
    
    def update_preset_name():        
        initial_value = current_preset
        new_value = preset_to_edit.value

        if new_value != initial_value and new_value != "":
            exist_preset = update_preset_exist_in_json(new_value)
            
            if exist_preset == True:
                preset_to_edit.error_text = 'Esse preset já existe'
                page.update()
                return
            else:
                update_preset_in_json(initial_value, new_value)
                preset_to_edit.error_text = None
            
            page.close(edit_preset_modal)
            create_presets_names()
            page.update()

        else:
            print("O novo nome da coluna não pode ser vazio ou igual ao nome inicial.")
    
    def preset_modal_action(e):
        action = e.control.key
        
        if action == 'delete':
            confirm_action_to_del_preset()
        else:
            update_preset_name()
            
        
    def handle_preset_edit_change(e):
        global value_input
        initial_value = current_preset
        value_input = e.control.value

        if value_input != initial_value and value_input != "":
            alter_preset_name_button.disabled = False
            alter_preset_name_button.opacity = 1.0
        else:
            alter_preset_name_button.disabled = True
            alter_preset_name_button.opacity = 0.4

        page.update()
    
    preset_to_edit = ft.TextField(
        label="Novo nome",
        border_color=ft.colors.OUTLINE_VARIANT,
        on_change=handle_preset_edit_change,
    )
    
    delete_preset_button = ft.Container(
        content=ft.Text(
            "Deletar",
            color=ft.colors.RED_400,
            weight=ft.FontWeight.W_500,
        ),
        key='delete',
        expand=1,
        padding=ft.padding.all(10),
        border_radius=ft.border_radius.all(10),
        border=ft.border.all(1, ft.colors.RED_400),
        on_click=preset_modal_action,
        alignment=ft.alignment.center,
    )
    alter_preset_name_button = ft.Container(
        content=ft.Text(
            "Modificar",
            color=ft.colors.OUTLINE,
            weight=ft.FontWeight.W_500,
        ),
        key='update',
        expand=1,
        padding=ft.padding.all(10),
        border_radius=ft.border_radius.all(10),
        border=ft.border.all(1, ft.colors.OUTLINE),
        on_click=preset_modal_action,
        alignment=ft.alignment.center,
        disabled=True,
        opacity=0.4,
    )
    
    delete_preset_button = [
        ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        delete_preset_button,
                        alter_preset_name_button,
                    ],
                    spacing=10,
                    width=initial_width,
                ),
                ft.TextButton("Cancelar", width=initial_width, on_click=lambda _: page.close(edit_preset_modal)),
            ],
            width=initial_width,
        ),
    ]
    
    edit_preset_modal = ft.AlertDialog(
        modal=True,
        title=ft.Row(
            controls=[
                ft.Text(
                    "Editar preset",
                    text_align=ft.TextAlign.CENTER,
                    size=20,
                ),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
        ),
        content=preset_to_edit,
        actions=delete_preset_button,
    )
    
    def open_modal_edit_preset(preset, e=None):
        global current_preset
        
        current_preset = preset
        
        preset_to_edit.value = preset
        preset_to_edit.error_text = None
        edit_preset_modal.content = preset_to_edit
        edit_preset_modal.actions = delete_preset_button
        
        page.open(edit_preset_modal)
        page.update()

    def new_preset(e):
        add_new_preset()
        create_presets_names()

    def on_preset_name_click(index, preset, e):
        global current_preset_column
        current_preset_column = preset
        
        for i, container in enumerate(presets_controls):
            if i == index - 1:
                container.bgcolor = ft.colors.OUTLINE_VARIANT
            else:
                container.bgcolor = ft.colors.TRANSPARENT
        
        create_json_elements(False, preset)
        page.update()

    def presets_names_create(presets_names_value):
        global presets_controls, first_preset
        presets_controls = []
        
        first_preset = None
        for key, value in presets_names_value.items():
            if isinstance(value, dict) and "Columns" in value:
                first_preset = key
                break
        
        for i, preset in enumerate(presets_names_value):
            index = i + 1
            checkbox = ft.Row(
                controls=[
                    ft.Text(
                        value=f'{preset}',
                        no_wrap=True
                    ),
                    ft.IconButton(
                        tooltip=f"Editar preset {index}",
                        icon=ft.icons.EDIT,
                        icon_size=15,
                        data={"index": i, "name": preset},
                        width=30,
                        on_click=partial(open_modal_edit_preset, preset),
                        style=ft.ButtonStyle(
                            padding=ft.padding.all(0),
                        ),
                    ),
                ],
                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                height=30,
            )
            checkbox_container = ft.Container(
                content=checkbox,
                padding=ft.Padding(top=5, bottom=9, left=20, right=11),
                border=ft.border.all(1, ft.colors.OUTLINE_VARIANT),
                border_radius=ft.border_radius.all(7),
                data=index,
                on_click=partial(on_preset_name_click, index, preset),
            )
            presets_controls.append(checkbox_container)

        add_preset = ft.IconButton(
            tooltip=f"Novo preset",
            icon=ft.icons.ADD,
            icon_size=20,
            on_click=new_preset,
            style=ft.ButtonStyle(
                padding=ft.padding.all(0),
            ),
        )
        presets_controls.append(add_preset)

        presets_names_area.controls = presets_controls
        
        if first_preset:
            on_preset_name_click(1, first_preset, None)
        else:
            print("Nenhum preset válido encontrado.")
            
        page.update()

    presets_names_area = ft.Row(
        width=initial_width,
        height=80,
        spacing=10,
        scroll="auto",
    )

    presets_row = ft.Row(
        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
        width=initial_width,
        spacing=10,
        wrap=True,
    )

    presets_column = ft.Column(
        visible=False,
        controls=[
            import_and_export,
            presets_names_area,
            edit_preset,
            presets_row,
        ]
    )

    def delete_column(e):
        data = e.control.data
        column_name_to_delete = data['name']

        remove_column_from_json(current_preset_column, column_name_to_delete)
        create_json_elements(False, current_preset_column)
        page.update()
        
    presets = []

    def presets_inputs_create(presets_value):
        presets.clear()
        for i, column in enumerate(presets_value):
            index = i+1
            checkbox = ft.Row(
                controls=[
                    ft.Text(
                        value=f'{index} - {column}',
                        width=176,
                        no_wrap=True
                    ),
                    ft.IconButton(
                        tooltip=f"Deletar coluna {index}",
                        icon=ft.icons.DELETE_FOREVER,
                        icon_color=ft.colors.RED,
                        icon_size=20,
                        data={"index": i, "name": column},
                        width=30,
                        on_click=delete_column,
                        style=ft.ButtonStyle(
                            padding=ft.padding.all(0),
                        ),
                    ),
                ],
                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                width=210,
                height=30,
            )
            checkbox_container = ft.Container(
                content=checkbox,
                padding=ft.padding.all(8),
                border=ft.border.all(1, ft.colors.WHITE10),
                border_radius=ft.border_radius.all(7),
                data=index,
                shadow=ft.BoxShadow(
                    spread_radius=1,
                    blur_radius=10,
                    color='#7F8C8D',
                    offset=ft.Offset(0, 0),
                    blur_style=ft.ShadowBlurStyle.OUTER,
                ),
            )
            presets.append(checkbox_container)
        presets_row.controls = presets
        page.update()
        
    progress_status_bar = ft.Container(
        margin=ft.Margin(top=15, bottom=0, left=0, right=0),
        content=ft.ProgressBar(
            border_radius=ft.border_radius.all(0),
            bgcolor=ft.colors.TRANSPARENT,
            width=initial_width,
            color="#eeeeee",
        )
    )
        
    async def delete_columns_preset(e):
        nonlocal selected_sheet_to_preset, modal_preset, status_text

        keeps = keeps_value
        selected_sheet = selected_sheet_to_preset
        selected_preset = selected_preset_value
        file_path = file_path_value
        
        status = ft.Row(
            spacing=10,
            wrap=True,
            controls=[
                status_text,
                progress_status_bar,
            ]
        )

        modal_preset.title = ft.Text("Preparando...", size=20)
        modal_preset.content = status
        modal_preset.actions = []
        page.update()

        async def process_deletion():
            status_text.value = f"Aguarde, estamos processando os dados..."
            page.update()

            await asyncio.sleep(1)

            status_msg, cols_to_remove, found_msg = delete_columns_with_preset(file_path, selected_sheet, selected_preset, keeps)

            status_text.value = "Quase lá..."
            page.update()
            
            await asyncio.sleep(2)

            if cols_to_remove:
                status_text.value = found_msg
                page.update()
                
                await asyncio.sleep(4)
            else:
                status_text.value = status_msg

            page.update()

        await process_deletion()
        page.close(modal_preset)

    def keeps_confirmation_value(e):
        global keeps_value

        keeps_value = e.control.value
        continue_button_modal_preset.disabled = False

        page.update()

    status_text = ft.Text(
        text_align=ft.TextAlign.CENTER,
        weight="bold",
        size=13,
    )

    confirmation_mensage_text = ft.Text(
        'Você deseja que as colunas do preset sejam mantidas ou deletadas?',
        text_align=ft.TextAlign.CENTER,
    )

    keeps_confirmation = ft.RadioGroup(
        content=ft.Row(
            controls=[
                ft.Radio(value='true', label="Manter colunas"),
                ft.Radio(value='false', label="Deletar colunas"),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
        ),
        on_change=keeps_confirmation_value,
    )

    confirmation_mensage = ft.Row(
        wrap=True,
        visible=False,
        controls={
            confirmation_mensage_text,
            keeps_confirmation,
        }
    )
    
    modal_preset = ft.AlertDialog(
        modal=True
    )

    continue_button_modal_preset = ft.TextButton(
        "Seguir",
        disabled=True,
        on_click=delete_columns_preset,
    )
    
    def selected_preset(e):
        global selected_preset_value
        
        selected_preset_value = e.control.value
        preset = load_json_data()

        if selected_preset_value in preset:
            preset_columns = preset[selected_preset_value].get('Columns', [])
        
        preset_columns_quantity = len(preset_columns)
        
        confirmation_mensage.visible = True 
        # confirmation_mensage_text.value = f'Quando clicado em "Seguir", serão removidas {preset_columns_quantity} colunas dessa planilha.'
        
        page.update()
    
    preset_dropdown_available = ft.Dropdown(
        label="Selecione o preset",
        border_color=ft.colors.OUTLINE_VARIANT,
        on_change=selected_preset,
        text_size=14,
    )
    
    preset_contain = ft.Container(
        visible=False,
        padding=ft.padding.all(0),
        content=preset_dropdown_available,
        border_radius=ft.border_radius.all(4),
        margin=ft.Margin(top=10, bottom=15, left=0, right=0),
        shadow=ft.BoxShadow(
            spread_radius=1,
            blur_radius=10,
            color='#050505',
            offset=ft.Offset(0, 1),
            blur_style=ft.ShadowBlurStyle.OUTER,
        ),
    )

    def select_preset():
        global presets_available
        
        presets_available = []
        presets_value = get_presets_names()
        
        for i, preset in enumerate(presets_value):
            preset_container = ft.dropdown.Option(
                text=preset,
                key=preset,
            )
            presets_available.append(preset_container)
            
        preset_dropdown_available.options = presets_available
        preset_contain.visible = True


    selected_sheet_to_preset = None

    def preset_tratment(e: ft.FilePickerResultEvent):
        global file_path_value

        preset = load_json_data()
        preset_columns = preset.get('Columns', [])
        
        SUPPORTED_EXTENSIONS = ['xlsx', 'xls']
        
        preset_dropdown_available.value = ''
        preset_contain.visible = False

        if e.files:
            selected_file = e.files[0]
            file_name_value = selected_file.name
            file_path_value = selected_file.path
            ext = file_name_value.split('.')[-1].lower()

            if ext in SUPPORTED_EXTENSIONS:
                sheetnames, columns_sheet = ftm.file_treatment(selected_file)
                
                def on_container_click(index):
                    nonlocal selected_sheet_to_preset
                    selected_sheet_to_preset = sheetnames[index]

                    for i, container in enumerate(radio_containers):
                        if i == index:
                            container.bgcolor = ft.colors.OUTLINE_VARIANT
                        else:
                            container.bgcolor = ft.colors.TRANSPARENT

                    select_preset()
                    page.update()

                radio_containers = [
                    ft.Container(
                        content=ft.Row(
                            controls=[
                                ft.Text(
                                    value=f"Planilha: {sheet}",
                                    size=14,
                                ),
                                ft.Text(
                                    value=f"Colunas: {columns_sheet.get(sheet, 0)}",
                                    size=11,
                                )
                            ],
                            spacing=5,
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        ),
                        border=ft.border.all(1, ft.colors.OUTLINE_VARIANT),
                        border_radius=ft.border_radius.all(7),
                        padding=ft.padding.all(10),
                        width=initial_width,
                        on_click=lambda _, index=i: on_container_click(index),
                    ) for i, sheet in enumerate(sheetnames)
                ]

                radio_group = ft.Row(
                    controls=[
                        *radio_containers,
                        preset_contain,
                        confirmation_mensage,
                    ],
                    width=initial_width,
                    spacing=10,
                    wrap=True,
                    alignment=ft.MainAxisAlignment.CENTER
                )

                nonlocal modal_preset
                
                modal_preset.title=ft.Text("Planilhas Disponíveis", size=20)
                modal_preset.actions=[
                    ft.Row(
                        controls=[
                            ft.TextButton("Cancelar", on_click=lambda _: page.close(modal_preset)),
                            continue_button_modal_preset,
                        ],
                        width=initial_width,
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    ),
                ]
                
                modal_preset.content = radio_group
                page.open(modal_preset)
                continue_button_modal_preset.disabled = True
                confirmation_mensage.visible = False

                if sheetnames is None:
                    result_title.value = 'Erro:'
                    file_path_text.value = 'Erro ao processar arquivo.'
            else:
                result_title.value = 'Erro:'
                file_path_text.value = 'Apenas arquivos de planilha (Excel) são permitidos.'

            page.update()

    file_to_preset = ft.FilePicker(on_result=preset_tratment)
    page.overlay.append(file_to_preset)

    def create_json_elements(first, preset_name):
        preset = load_json_data()
        
        if first == True:
            preset_name = first_preset

        if preset_name in preset:
            preset_columns = preset[preset_name].get('Columns', [])
        else:
            preset_columns = []

        presets_inputs_create(preset_columns)
        
    def create_presets_names():
        presets_names = get_presets_names()
        presets_names_create(presets_names)
    
    def presets_zone(e):
        nonlocal presets_column
        
        ensure_documents_json_file()

        if e.control.value == True:
            create_presets_names()
            create_json_elements(True, None)

        presets_column.visible = e.control.value
        page.update()

    presets_area = ft.Container(
        visible=False,
        padding=ft.Padding(top=10, bottom=20, left=20, right=20),
        width=initial_width,
        content=ft.Column(
            controls=[
                ft.Row(
                    alignment=ft.MainAxisAlignment.CENTER,
                    controls=[EQS_title_value]
                ),
                ft.Row(
                    controls=[
                        ft.Container(
                            content=ft.ElevatedButton(
                                text="Iniciar",
                                on_click=lambda _: file_to_preset.pick_files(),
                            ),
                            expand=1,
                        ),
                        ft.Container(
                            content=ft.Row(
                                controls=[
                                    ft.Checkbox(
                                        label='Editar preset',
                                        value=False,
                                        on_change=presets_zone,
                                    )
                                ],
                                alignment=ft.MainAxisAlignment.CENTER,
                            ),
                            border=ft.border.all(1, ft.colors.OUTLINE_VARIANT),
                            border_radius=ft.border_radius.all(50),
                            expand=1,
                        ),
                    ],
                    spacing=10,
                ),
                presets_column,
            ]
        )
    )

    def changed_tab(e):
        my_index = e.control.selected_index
        home_area_route.visible = True if my_index == 0 else False
        presets_area.visible = True if my_index == 1 else False
        page.update()
    
    page.navigation_bar = ft.NavigationBar(
        border=ft.border.only(top=ft.border.BorderSide(1, "black")),
        on_change=changed_tab,
        selected_index=0,
        destinations= [
            ft.NavigationBarDestination(icon=ft.icons.HOME, label='Inicio'),
            ft.NavigationBarDestination(icon=ft.icons.WORK, label='Presets'),
        ]
    )

    page.add(
        title,
        home_area_route,
        presets_area,
    )

    if new_version == True:
        page.open(banner_new_vesion)

    page.update()